
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.contrib import messages, auth
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.shortcuts import render, redirect
from django.views import View
import json
from django.contrib.auth.models import User
from django.http import JsonResponse
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, LineChart, Reference
from io import BytesIO
import calendar
from dateutil import parser
from django.db.models import Count
from django.db.models.functions import TruncDate


from django.contrib.auth.decorators import login_required


def Login(request):
    if request.method == 'POST':
        usern = request.POST['usern']
        passw = request.POST['passw']

        user = auth.authenticate(username=usern, password=passw)
        if user is not None:
            auth.login(request, user)
            messages.success(request, 'You are now logged in ' + usern)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid Credentials')
            return redirect('login')
    else:
        return render(request, 'login.html')


# Dashboard Module
@login_required(login_url="")
def Dashboard(request):
    labels = []
    data = []
    lbls = []
    dt = []

    queryset = Tracking.objects.order_by('-Overspeeding')[:5]
    for d in queryset:
        labels.append(d.Driver.name)
        data.append(d.Overspeeding)

    queryset = Fuel.objects.order_by('-Liters_taken')[:5]
    for f in queryset:
        lbls.append(f.Vehicle.Plate_no)
        dt.append(f.Liters_taken)
    return render(request, 'index.html', {
        'labels': labels,
        'data': data,
        'lbls': lbls,
        'dt': dt,
    })


def Logout(request):
    if request.method == 'POST':
        auth.logout(request)
        messages.success(request, 'You are now logged out')
        return redirect('Login')
    return redirect('Login')


@login_required(login_url="")
def journey(request):
    if request.method == 'POST':
        date = request.POST['date']
        driver = request.POST['nam']
        vehicles = request.POST['vehicle']
        stop = request.POST['stop']
        vehicle_condition = request.POST['vehicle_condition']
        initial_jmp = request.POST['initial_jmp']
        final_jmp = request.POST['final_jmp']
        start_trip = request.POST['start_trip']
        destination = request.POST['destination']
        stop_trip = request.POST['stop_trip']
        trip_approved = request.POST['trip_approved']
        passengers = request.POST['passengers']
        veh = Vehicle.objects.get(Plate_no=vehicles)
        dr = Driver.objects.get(name=driver)
        app = Aprover.objects.get(name=trip_approved)
        start = veh.odometer
        veh.odometer = int(stop)
        change = int(stop) - start

        if int(stop) <= start:
            vehicles = Vehicle.objects.all()
            drivers = Driver.objects.all()
            aprovers = Aprover.objects.all()
            return render(request, 'addjmp.html', {"vehicles": vehicles, "drivers": drivers, "aprovers": aprovers, "message": "The odometer reading you entered is less or equal to the existing odometer reading"})

        instance = Journey(Date=date, Driver=dr, Vehicle=veh, Start_odometer_reading=start, Stop_odometer_reading=stop, Round_trip_distance=change, vehicle_condition=vehicle_condition,
                           initial_JMP=initial_jmp, final_JMP=final_jmp, start_trip=start_trip, stop_trip=stop_trip, destination=destination, Approver=app, passengers=passengers)

        dist = veh.Distance_remaining - change
        if dist < 0:
            veh.Distance_remaining = 0
        else:
            veh.Distance_remaining = dist

        try:
            veh.save()
            instance.save()
            vehicles = Vehicle.objects.all()
            drivers = Driver.objects.all()
            aprovers = Aprover.objects.all()
            return render(request, 'addjmp.html', {"vehicles": vehicles, "drivers": drivers, "aprovers": aprovers, "message": "Saved Successfully"})
        except:
            vehicles = Vehicle.objects.all()
            drivers = Driver.objects.all()
            aprovers = Aprover.objects.all()
            return render(request, 'addjmp.html', {"vehicles": vehicles, "drivers": drivers, "aprovers": aprovers, "message": "Error Occured"})
    else:
        vehicles = Vehicle.objects.all()
        drivers = Driver.objects.all()
        aprovers = Aprover.objects.all()
        return render(request, 'addjmp.html', {"vehicles": vehicles, "drivers": drivers, "aprovers": aprovers})


@login_required(login_url="")
def vehicle(request):
    vehicles = Vehicle.objects.all()
    objects = []
    for veh in vehicles:
        if veh.Distance_remaining < 500:
            objects.append({"vehicle": veh, "status": "Needs servicing"})
        else:
            objects.append({"vehicle": veh, "status": "Good"})
    return render(request, "vehiclemanagement.html", {"objects": objects})


@login_required(login_url="")
def fuel(request):
    if request.method == "POST":
        date = request.POST['date']
        vehicles = request.POST['vehicle']
        driver = request.POST['driver']
        current_odometer_reading = request.POST['current_odometer_reading']
        liters_taken = request.POST['liters_taken']
        cost = request.POST['cost']
        receipt = request.POST['receipt']
        veh = Fuel.objects.filter(Vehicle__Plate_no=vehicles)
        dr = Driver.objects.get(name=driver)
        dte = []
        vehic = Vehicle.objects.get(Plate_no=vehicles)
        if veh:
            for i in veh:
                dte.append(i.Date)
            vehi = Fuel.objects.get(Vehicle__Plate_no=vehicles, Date=max(dte))
            tot = int(current_odometer_reading) - \
                int(vehi.current_odometer_reading)
            eff = tot / vehi.Liters_taken
            vari = vehic.expected_efficiency - eff
        else:
            eff = vehic.expected_efficiency
            tot = 0
            vari = 0
        vehicles = Vehicle.objects.all()
        drivers = Driver.objects.all()
        fuels = Fuel.objects.all()
        instance = Fuel(Date=date, Vehicle=vehic, Driver=dr, current_odometer_reading=current_odometer_reading,
                        total_distance_from_previous_fueling=tot, Liters_taken=liters_taken, cost=cost, receipt=receipt, effiency=eff, variation=vari)
        try:
            instance.save()
            vehicles = Vehicle.objects.all()
            drivers = Driver.objects.all()
            fuels = Fuel.objects.all()
            return render(request, "fuel.html", {"fuels": fuels, "vehicles": vehicles, "drivers": drivers, "message": "Saved Successfully"})
        except:
            vehicles = Vehicle.objects.all()
            drivers = Driver.objects.all()
            fuels = Fuel.objects.all()
            return render(request, "fuel.html", {"fuels": fuels, "vehicles": vehicles, "drivers": drivers, "message": "Error Occured"})
    else:
        vehicles = Vehicle.objects.all()
        drivers = Driver.objects.all()
        fuels = Fuel.objects.all()
        return render(request, "fuel.html", {"fuels": fuels, "vehicles": vehicles, "drivers": drivers})

# Create your views here.


@login_required(login_url="")
def tracking(request):
    if request.method == "POST":
        vehicles = Vehicle.objects.all()
        drivers = Driver.objects.all()
        vehicle = request.POST['vehicle']
        driver = request.POST['driver']
        date = request.POST['date']
        daily_distance = request.POST['distance']
        overspeeding = request.POST['speeding']
        jmp_daily_distance = request.POST['jmpdistance']
        veh = Vehicle.objects.get(Plate_no=vehicle)
        dr = Driver.objects.get(name=driver)
        instance = Tracking(Vehicle=veh, Driver=dr, Date=date, Daily_vehicle_tracking_distance=daily_distance,
                            Overspeeding=overspeeding, JMP_daily_distance=jmp_daily_distance)
        try:
            instance.save()
            return render(request, 'tracking.html', {"vehicles": vehicles, "drivers": drivers, "message": "Saved Successfully"})
        except:
            return render(request, 'tracking.html', {"vehicles": vehicles, "drivers": drivers, "message": "Unable to  save"})
    else:
        vehicles = Vehicle.objects.all()
        drivers = Driver.objects.all()
        return render(request, 'tracking.html', {"vehicles": vehicles, "drivers": drivers})


def vehiclemaintenance(request):
    return render(request, 'vehiclemaintenance.html')


def vehicleinspection(request):
    vehicles = Vehicle.objects.all()
    if request.method == "POST":
        vehicle = request.POST['vehicle']
        inspection_date = request.POST['inspection_date']
        next_inspection_date = request.POST['next_inspection_date']
        vhe = Vehicle.objects.get(Plate_no=vehicle)

        vhe.last_inspection_date = datetime.strptime(
            inspection_date, '%Y-%m-%d')
        vhe.next_inspection_date = datetime.strptime(
            next_inspection_date, '%Y-%m-%d')
        try:
            vhe.save()
            return render(request, 'addvehicleinspection.html', {"vehicles": vehicles, "message": "Saved Successfully"})
        except:
            return render(request, 'addvehicleinspection.html', {"vehicles": vehicles, "message": "Error Occured"})
    else:
        return render(request, 'addvehicleinspection.html', {"vehicles": vehicles})


def vehicleinsuarance(request):
    vehicles = Vehicle.objects.all()
    if request.method == "POST":
        vehicle = request.POST['vehicle']
        insuarance_date = request.POST['insuarance_date']
        insuarance_expiry = request.POST['insuarance_expiry']
        car = Vehicle.objects.get(Plate_no=vehicle)

        car.Insuarance_start_date = datetime.strptime(
            insuarance_date, '%Y-%m-%d')
        car.Insuarance_expiry_date = datetime.strptime(
            insuarance_expiry, '%Y-%m-%d')
        try:
            car.save()
            return render(request, 'addvehicleinsuarance.html', {"vehicles": vehicles, "message": "Saved Successfully"})
        except:
            return render(request, 'addvehicleinsuarance.html', {"vehicles": vehicles, "message": "Error Occured"})
    else:
        return render(request, 'addvehicleinsuarance.html', {"vehicles": vehicles})


def vehicleservicing(request):
    vehicles = Vehicle.objects.all()
    if request.method == "POST":
        vehicle = request.POST['vehicle']
        service_distance = request.POST['service_distance']
        cre = Vehicle.objects.get(Plate_no=vehicle)

        cre.last_service_distance = int(service_distance)
        cre.Distance_remaining = 5000
        try:
            cre.save()
            return render(request, 'addvehicleservice.html', {"vehicles": vehicles, "message": "Saved Successfully"})
        except:
            return render(request, 'addvehicleservice.html', {"vehicles": vehicles, "message": "Error Occured"})
    else:
        return render(request, 'addvehicleservice.html', {"vehicles": vehicles})


def fuelingreport(request):
    if request.method == 'POST':
        month = request.POST['month']
        year = request.POST['year']

        value = {
            'January': 1,
            'February': 2,
            'March': 3,
            'April': 4,
            'May': 5,
            'June': 6,
            'July': 7,
            'August': 8,
            'September': 9,
            'October': 10,
            'November': 11,
            'December': 12
        }
        month = value[month]

        print("Month:", month)
        print("Year:", year)

        data_queryset = Fuel.objects.filter(Date__month=month, Date__year=year)

        wb = Workbook()
        ws = wb.active
        logo = Image('static/Techminds-logo.png')

        ws.merge_cells('A1:Z1')
        logo_cell = ws.cell(row=1, column=6)
        logo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.add_image(logo, 'F1')

        ws.row_dimensions[1].height = 50

        ws.append(['DATE', 'VEHICLE', 'DRIVER', 'ODOMETER READING',
                  'TOTAL DISTANCE COVERED', 'LITERS', 'COST', 'EFFICIENCY', 'VARIATION'])
        for cell in ws[2]:
            cell.fill = PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid')

        total_liters = 0
        total_cost = 0

        for item in data_queryset:
            ws.append([item.Date, item.Vehicle.Plate_no, item.Driver.name, item.current_odometer_reading,
                      item.total_distance_from_previous_fueling, item.Liters_taken, item.cost, item.effiency, item.variation])
            total_liters += item.Liters_taken
            total_cost += item.cost

        ws.append(['Total Liters', total_liters])
        ws.append(['Total Cost', total_cost])

        content_font = Font(size=12)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=fuelingreport_{month}_{year}.xlsx'
        wb.save(response)
        return response
    else:
        return render(request, 'fuelingreport.html')


def jmpreports(request):
    if request.method == 'POST':
        month = request.POST['month']
        year = request.POST['year']

        value = {
            'January': 1,
            'February': 2,
            'March': 3,
            'April': 4,
            'May': 5,
            'June': 6,
            'July': 7,
            'August': 8,
            'September': 9,
            'October': 10,
            'November': 11,
            'December': 12
        }
        month = value[month]

        print("Month:", month)
        print("Year:", year)

        data_queryset = Journey.objects.filter(
            Date__month=month, Date__year=year)

        wb = Workbook()
        ws = wb.active

        logo = Image('static/Techminds-logo.png')

        ws.merge_cells('A1:Z1')
        logo_cell = ws.cell(row=1, column=6)
        logo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.add_image(logo, 'F1')

        ws.row_dimensions[1].height = 50

        ws.append(['DATE',  'DRIVER', 'VEHICLE', 'START', 'STOP',
                   'INITIAL JMP', 'FINAL JMP', 'ROUND TRIP DISTANCE', 'VEHICLE CONDITION', 'START TRIP', 'DESTINATION', 'STOP TRIP', 'TRIP APPROVED', 'PASSENGERS'])

        for cell in ws[2]:
            cell.fill = PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid')

        for item in data_queryset:
            ws.append([item.Date,  item.Driver.name, item.Vehicle.Plate_no, item.Start_odometer_reading,
                       item.Stop_odometer_reading, item.initial_JMP, item.final_JMP, item.Round_trip_distance, item.vehicle_condition, item.start_trip,  item.destination, item.stop_trip, item.Approver.name, item.passengers])

        content_font = Font(size=12)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=jmpreport_{month}_{year}.xlsx'
        wb.save(response)
        return response
    else:
        return render(request, 'jmpreports.html')


def trackingreports(request):
    if request.method == 'POST':
        month = request.POST['month']
        year = request.POST['year']

        value = {
            'January': 1,
            'February': 2,
            'March': 3,
            'April': 4,
            'May': 5,
            'June': 6,
            'July': 7,
            'August': 8,
            'September': 9,
            'October': 10,
            'November': 11,
            'December': 12
        }
        month = value[month]

        print("Month:", month)
        print("Year:", year)

        data_queryset = Tracking.objects.filter(
            Date__month=month, Date__year=year)

        overspeeding_data = {}

        for item in data_queryset:
            driver_name = item.Driver.name
            if driver_name in overspeeding_data:
                overspeeding_data[driver_name] += item.Overspeeding
            else:
                overspeeding_data[driver_name] = item.Overspeeding

        wb = Workbook()
        ws = wb.active

        logo = Image('static/Techminds-logo.png')

        ws.merge_cells('A1:Z1')
        logo_cell = ws.cell(row=1, column=6)
        logo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.add_image(logo, 'F1')

        ws.row_dimensions[1].height = 50

        ws.append(['DATE',  'DRIVER', 'VEHICLE', 'TRACKING DISTANCE',
                  'OVERSPEEDING', 'JMP DAILY DISTANCE'])

        for cell in ws[2]:
            cell.fill = PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid')

        for item in data_queryset:
            ws.append([item.Date, item.Driver.name, item.Vehicle.Plate_no,
                      item.Daily_vehicle_tracking_distance, item.Overspeeding, item.JMP_daily_distance])

        content_font = Font(size=12)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=trackingreport_{month}_{year}.xlsx'

        wb.save(response)
        return response
    else:
        return render(request, 'trackingreports.html')
